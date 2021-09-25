import os

import openpyxl
from loguru import logger

import db_connect

logger.add(
    "bot_debug.log",
    format="{time} {level} {message}",
    level="DEBUG",
    rotation="10 MB",
    retention="7 days",
    compression="zip",
)

wb_with_table = openpyxl.load_workbook(
    os.path.join(r"C:\Dev\encosts_test\upload_names", "названия точек.xlsm") #Захаркодил путь, тк без него excel ищет файл в других каталогах.
)
work_sheet = wb_with_table.active
title = work_sheet.cell(row=1, column=2).value


@logger.catch
def get_cell_value(index, column_num):
    """
    Функция возвращает значения, полученные из таблицы excel.
    """
    return work_sheet.cell(row=index, column=column_num).value


@logger.catch
def get_new_data():
    """
    Функция отбирает данные из таблицы и возвращает данные для добавления
    в таблицу.
    """
    rows_amount = work_sheet.max_row
    BAD_RESULT = [None, "-"]
    result = []
    for index in range(2, rows_amount):
        endpoint_id = get_cell_value(index, 1)
        endpoint_names = get_cell_value(index, 2)
        if endpoint_id not in BAD_RESULT and endpoint_names not in BAD_RESULT:
            result.append((endpoint_id, endpoint_names))
    endpoint_id = []
    endpoint_names = []
    for elem in result:
        endpoint_id.append(elem[0])
        endpoint_names.append(elem[1])
    return endpoint_id, endpoint_names


if __name__ == "__main__":
    db_connect.init_db()
